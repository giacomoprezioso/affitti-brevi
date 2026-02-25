"""
Generazione report e pivot dalle prenotazioni.

Legge i dati dal foglio 'elenco' dell'Excel e produce:
  - DataFrame pivot per Streamlit (st.dataframe)
  - Tabelle riepilogative per mese, proprietà, piattaforma
"""

import pandas as pd
from datetime import date
from typing import List, Optional

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import EXCEL_PATH, SHEET_ELENCO


def load_elenco(excel_path: str = None) -> pd.DataFrame:
    """
    Carica il foglio 'elenco' dall'Excel.
    Usa openpyxl con data_only=True per leggere valori (non formule).
    """
    if excel_path is None:
        excel_path = EXCEL_PATH

    from openpyxl import load_workbook
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[SHEET_ELENCO]

    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c else f"col_{j}" for j, c in enumerate(row)]
            continue
        if any(v is not None for v in row):
            rows.append(row)

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows, columns=headers)
    return df


def build_bookings_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Filtra e struttura le prenotazioni dal foglio elenco.
    Righe di tipo 'incasso' con importo positivo.
    """
    if df.empty:
        return pd.DataFrame()

    # Rinomina colonne in base alla posizione (più robusto dei nomi)
    col_names = list(df.columns)

    # Mapping posizione → nome logico
    col_map_names = {
        0: "proprieta", 1: "check_in", 2: "check_out", 3: "mese",
        4: "tax", 5: "importo", 6: "tipo", 7: "causale",
        8: "piattaforma", 9: "ospite", 10: "documento", 11: "codice",
        12: "data", 15: "notti", 17: "ritenuta", 18: "incassato",
        19: "lordo", 20: "commissione", 21: "payment_charge", 22: "vat",
    }

    df2 = pd.DataFrame()
    for idx, name in col_map_names.items():
        if idx < len(col_names):
            df2[name] = df.iloc[:, idx]

    # Filtra solo righe prenotazione (tipo incasso o causale da clienti)
    mask_booking = (
        df2["tipo"].astype(str).str.lower().str.contains("incasso", na=False) |
        df2["causale"].astype(str).str.lower().str.contains("da clienti", na=False)
    )
    df_bookings = df2[mask_booking].copy()

    if df_bookings.empty:
        return pd.DataFrame()

    # Conversioni tipi
    for col in ["importo", "lordo", "incassato", "ritenuta", "commissione", "notti"]:
        if col in df_bookings.columns:
            df_bookings[col] = pd.to_numeric(df_bookings[col], errors="coerce").fillna(0)

    # Proprietà come stringa leggibile
    df_bookings["proprieta"] = df_bookings["proprieta"].apply(
        lambda x: f"Caldiero {int(float(x))}" if pd.notna(x) and str(x) not in ("", "None") else "N/D"
    )

    # Mese come int
    df_bookings["mese"] = pd.to_numeric(df_bookings["mese"], errors="coerce").fillna(0).astype(int)

    # Data check-in (può essere numero seriale Excel o già date)
    def to_date_safe(val):
        if val is None or str(val) in ("", "None", "nan"):
            return None
        if isinstance(val, (date,)):
            return val
        try:
            import datetime as dt
            if isinstance(val, (int, float)):
                # Numero seriale Excel → data
                from openpyxl.utils.datetime import from_excel
                return from_excel(int(val))
            return pd.to_datetime(str(val)).date()
        except Exception:
            return None

    if "check_in" in df_bookings.columns:
        df_bookings["check_in"] = df_bookings["check_in"].apply(to_date_safe)
        df_bookings["anno_mese"] = df_bookings["check_in"].apply(
            lambda d: d.strftime("%Y-%m") if d else "N/D"
        )
    else:
        df_bookings["anno_mese"] = "N/D"

    return df_bookings


def pivot_by_month_property(df_bookings: pd.DataFrame) -> pd.DataFrame:
    """Pivot: mese × proprietà, valori lordo/netto/notti."""
    if df_bookings.empty:
        return pd.DataFrame()

    pivot = df_bookings.pivot_table(
        values=["lordo", "incassato", "notti"],
        index="anno_mese",
        columns="proprieta",
        aggfunc="sum",
        fill_value=0,
        margins=True,
        margins_name="TOTALE",
    )
    return pivot


def pivot_by_platform(df_bookings: pd.DataFrame) -> pd.DataFrame:
    """Riepilogo per piattaforma."""
    if df_bookings.empty:
        return pd.DataFrame()

    if "piattaforma" not in df_bookings.columns:
        return pd.DataFrame()

    summary = df_bookings.groupby("piattaforma").agg(
        prenotazioni=("codice", "count"),
        lordo_totale=("lordo", "sum"),
        netto_totale=("incassato", "sum"),
        notti_totali=("notti", "sum"),
    ).reset_index()

    return summary


def bookings_list(df_bookings: pd.DataFrame) -> pd.DataFrame:
    """Lista prenotazioni per visualizzazione tabellare."""
    if df_bookings.empty:
        return pd.DataFrame()

    cols = [c for c in ["anno_mese", "proprieta", "piattaforma", "ospite",
                         "check_in", "check_out", "notti", "lordo",
                         "commissione", "incassato", "codice"] if c in df_bookings.columns]
    return df_bookings[cols].sort_values("anno_mese", ascending=False)


def costs_summary(excel_path: str = None) -> pd.DataFrame:
    """Riepilogo costi (bollette, pulizie, ecc.)."""
    df = load_elenco(excel_path)
    if df.empty:
        return pd.DataFrame()

    col_names = list(df.columns)

    df2 = pd.DataFrame()
    df2["proprieta"] = df.iloc[:, 0]
    df2["mese"] = df.iloc[:, 3]
    df2["importo"] = pd.to_numeric(df.iloc[:, 5], errors="coerce").fillna(0)
    df2["tipo"] = df.iloc[:, 6].astype(str)
    df2["causale"] = df.iloc[:, 7].astype(str)
    df2["fornitore"] = df.iloc[:, 8].astype(str)

    # Solo costi (importo negativo, tipo ordinarie/straordinarie)
    mask_cost = df2["importo"] < 0
    df_costs = df2[mask_cost].copy()

    if df_costs.empty:
        return pd.DataFrame()

    df_costs["proprieta"] = df_costs["proprieta"].apply(
        lambda x: f"Caldiero {int(float(x))}" if pd.notna(x) and str(x) not in ("", "None") else "N/D"
    )

    summary = df_costs.groupby(["causale", "fornitore"]).agg(
        totale=("importo", "sum"),
        occorrenze=("importo", "count"),
    ).reset_index()
    summary["totale"] = summary["totale"].round(2)

    return summary
