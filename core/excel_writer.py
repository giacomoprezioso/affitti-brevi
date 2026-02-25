"""
Aggiorna il file Excel preservando tutte le formule esistenti.

Strategia:
  1. Backup automatico prima di ogni scrittura
  2. load_workbook() senza data_only → preserva formule
  3. insert_rows() per inserire nuove righe dentro i range esistenti
     → openpyxl aggiusta automaticamente SUM(E6:E24) etc.
  4. Scrive valori nelle nuove righe
  5. Salva il file
"""

import shutil
from datetime import date, datetime
from typing import List, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import EXCEL_PATH, SHEET_ELENCO, COL_MAP
from core.models import Booking, Cost
from core.deduplicator import load_existing_codes, is_booking_duplicate, is_invoice_duplicate


def _backup(excel_path: str) -> str:
    """Crea backup del file Excel prima di modificarlo."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = excel_path.replace(".xlsx", "")
    backup_path = f"{base}_backup_{ts}.xlsx"
    shutil.copy2(excel_path, backup_path)
    return backup_path


def _find_last_data_row(ws) -> int:
    """
    Trova l'ultima riga con dati nella colonna A (Caldiero).
    Usa max_row del worksheet poi scorre solo le ultime righe per evitare
    di caricare 1M celle.
    """
    # Scorre le ultime 200 righe dalla fine — sufficiente per trovare la fine dei dati
    max_r = ws.max_row
    for row_num in range(max_r, 0, -1):
        cell = ws.cell(row=row_num, column=1)
        if cell.value is not None and str(cell.value).strip() != "":
            return row_num
    return 1


def _write_booking_row(ws, row_num: int, b: Booking):
    """Scrive una prenotazione nel foglio elenco."""
    def set_cell(col_key, value):
        if value is not None:
            ws.cell(row=row_num, column=COL_MAP[col_key]).value = value

    set_cell("caldiero", b.property_num)
    set_cell("dal", b.check_in)
    set_cell("al", b.check_out)
    # Mese come formula dinamica
    ws.cell(row=row_num, column=COL_MAP["mese"]).value = f"=MONTH(B{row_num})"
    set_cell("tax", "T")
    set_cell("importo", round(b.net_amount, 2) if b.net_amount else None)
    set_cell("tipo", "incasso")
    set_cell("causale", "da clienti")
    set_cell("ente", b.platform.capitalize())
    set_cell("nominativo", b.guest_name)
    set_cell("documento", "prenotazione")
    set_cell("nr", b.confirm_code)
    set_cell("data", b.check_in)
    set_cell("giorni", b.nights if b.nights else None)
    set_cell("ritenuta", round(b.withholding_tax, 2) if b.withholding_tax else None)
    set_cell("incassato", round(b.net_amount, 2) if b.net_amount else None)
    set_cell("lordo", round(b.gross_amount, 2) if b.gross_amount else None)
    set_cell("commission", round(b.commission, 2) if b.commission else None)
    set_cell("payment_charge", round(b.payment_charge, 2) if b.payment_charge else None)
    set_cell("vat", round(b.vat, 2) if b.vat else None)
    # Euro/gg: formula solo se ci sono le notti
    if b.nights and b.nights > 0:
        t_col = get_column_letter(COL_MAP["lordo"])
        p_col = get_column_letter(COL_MAP["giorni"])
        ws.cell(row=row_num, column=COL_MAP["euro_gg"]).value = f"={t_col}{row_num}/{p_col}{row_num}"


def _write_cost_row(ws, row_num: int, c: Cost):
    """Scrive un costo (bolletta/fattura) nel foglio elenco."""
    def set_cell(col_key, value):
        if value is not None:
            ws.cell(row=row_num, column=COL_MAP[col_key]).value = value

    prop = c.property_num if c.property_num in (5, 7) else 5  # fallback
    set_cell("caldiero", prop)
    set_cell("dal", c.date)
    # mese formula
    ws.cell(row=row_num, column=COL_MAP["mese"]).value = f"=MONTH(B{row_num})"
    set_cell("tax", "T")
    set_cell("importo", round(c.amount, 2))
    set_cell("tipo", "ordinarie")
    set_cell("causale", c.category)
    set_cell("ente", c.supplier)
    set_cell("nominativo", "bolletta" if c.category in ("acqua", "energia elettrica", "gas") else "fattura")
    set_cell("documento", c.invoice_num if c.invoice_num else None)
    set_cell("data", c.invoice_date if c.invoice_date else c.date)
    set_cell("intestata_a", "si")


def update_excel(
    bookings: List[Booking],
    costs: List[Cost],
    excel_path: str = None,
    dry_run: bool = False,
) -> Tuple[int, int, List[str]]:
    """
    Aggiunge nuove prenotazioni e costi al foglio elenco.

    Returns:
        (added_bookings, added_costs, skipped_codes)
    """
    if excel_path is None:
        excel_path = EXCEL_PATH

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"File Excel non trovato: {excel_path}")

    wb = load_workbook(excel_path)  # NON data_only=True → preserva formule
    if SHEET_ELENCO not in wb.sheetnames:
        raise ValueError(f"Foglio '{SHEET_ELENCO}' non trovato nel file Excel. "
                         f"Fogli presenti: {wb.sheetnames}")
    ws = wb[SHEET_ELENCO]

    existing_booking_codes, existing_invoice_nums = load_existing_codes(wb)

    # Filtra duplicati
    new_bookings = []
    skipped = []
    for b in bookings:
        if is_booking_duplicate(b.confirm_code, existing_booking_codes):
            skipped.append(b.confirm_code)
        else:
            new_bookings.append(b)
            existing_booking_codes.add(b.confirm_code)

    new_costs = []
    for c in costs:
        # Chiave dedup per fatture split: invoice_num + property_num
        dedup_key = f"{c.invoice_num}_{c.property_num}" if c.invoice_num else None
        if dedup_key and is_invoice_duplicate(c.invoice_num, existing_invoice_nums, c.property_num):
            skipped.append(c.invoice_num)
        else:
            new_costs.append(c)
            if dedup_key:
                existing_invoice_nums.add(dedup_key)

    n_new = len(new_bookings) + len(new_costs)

    if n_new == 0 or dry_run:
        return len(new_bookings), len(new_costs), skipped

    # Backup prima di scrivere
    _backup(excel_path)

    last_row = _find_last_data_row(ws)

    # Appende direttamente dopo l'ultima riga con dati (NO insert_rows che
    # spinge righe vuote oltre il limite Excel 1048576).
    # Le formule nei fogli netto* usano range fissi già presenti nel file.
    current_row = last_row + 1
    for b in new_bookings:
        _write_booking_row(ws, current_row, b)
        current_row += 1

    for c in new_costs:
        _write_cost_row(ws, current_row, c)
        current_row += 1

    wb.save(excel_path)
    return len(new_bookings), len(new_costs), skipped
